import qrcode
import openpyxl as xl

wb = xl.load_workbook('AutoMail/database.xlsx')
sheet = wb['Sheet1']

for i in range(2, sheet.max_row+1):
    data = sheet.cell(i, 1).value
    qr = qrcode.QRCode(version=1, box_size=10, border=1)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#d81d7e', back_color='#ece4fb')
    # img = qr.make_image(fill_color='black', back_color='white')
    img.save(f'ISIMails/qrimages/{sheet.cell(i, 1).value}.png')
